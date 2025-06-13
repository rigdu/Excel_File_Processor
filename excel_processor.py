import tkinter as tk
from tkinter import filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
import openpyxl
import os

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Processor")
        self.root.geometry("600x400")

        # Variables for input fields
        self.product_code_var = tk.StringVar()
        self.product_name_var = tk.StringVar()
        self.bc_var = tk.StringVar()
        self.dia_var = tk.StringVar()
        self.replicate_var = tk.BooleanVar()

        # GUI Elements
        self.create_gui()

        # File path storage
        self.file_path = None

    def create_gui(self):
        # Upload/Drag-and-Drop Button
        self.upload_label = tk.Label(self.root, text="Drag and Drop or Click to Upload .xlsx File")
        self.upload_label.pack(pady=10)

        self.upload_button = tk.Button(self.root, text="Upload Excel File", command=self.upload_file)
        self.upload_button.pack(pady=5)
        self.upload_button.drop_target_register(DND_FILES)
        self.upload_button.dnd_bind('<<Drop>>', self.drop_file)

        # Input Fields
        tk.Label(self.root, text="Product Code (Column A):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.product_code_var).pack()

        tk.Label(self.root, text="Product Name (Column B):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.product_name_var).pack()

        tk.Label(self.root, text="BC (Column C):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.bc_var).pack()

        tk.Label(self.root, text="Dia (Column D):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.dia_var).pack()

        # Replicate Toggle
        tk.Checkbutton(self.root, text="Replicate Data (Double Rows)", variable=self.replicate_var).pack(pady=10)

        # Process and Save Button
        tk.Button(self.root, text="Process and Save", command=self.process_and_save).pack(pady=20)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.file_path = file_path
            self.upload_label.config(text=f"Selected: {os.path.basename(file_path)}")

    def drop_file(self, event):
        self.file_path = event.data
        if self.file_path.startswith('{'):
            self.file_path = self.file_path.strip('{}')
        self.upload_label.config(text=f"Selected: {os.path.basename(self.file_path)}")

    def process_and_save(self):
        if not self.file_path or not self.file_path.endswith('.xlsx'):
            messagebox.showerror("Error", "Please upload a valid .xlsx file")
            return

        if not all([self.product_code_var.get(), self.product_name_var.get(), self.bc_var.get(), self.dia_var.get()]):
            messagebox.showerror("Error", "Please fill all input fields")
            return

        try:
            # Load workbook
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            # Read data from A, B, C
            data = []
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)

            # Clear original data in A, B, C, D, E, F, G, H
            for row in range(2, sheet.max_row + 1):
                for col in range(1, 9):  # Columns A to H
                    sheet.cell(row=row, column=col).value = None

            # Shift data to E, F, G
            for i, (col_a, col_b, col_c) in enumerate(data, start=2):
                sheet.cell(row=i, column=5).value = col_a  # E
                sheet.cell(row=i, column=6).value = col_b  # F
                sheet.cell(row=i, column=7).value = col_c  # G

            # Fill A, B, C, D with user input
            num_rows = len(data)
            for i in range(2, num_rows + 2):
                sheet.cell(row=i, column=1).value = self.product_code_var.get()  # A
                sheet.cell(row=i, column=2).value = self.product_name_var.get()  # B
                sheet.cell(row=i, column=3).value = self.bc_var.get()  # C
                sheet.cell(row=i, column=4).value = self.dia_var.get()  # D

            # Handle replication if toggle is checked
            if self.replicate_var.get():
                for i, row_data in enumerate(data, start=num_rows + 2):
                    sheet.cell(row=i, column=1).value = self.product_code_var.get()  # A
                    sheet.cell(row=i, column=2).value = self.product_name_var.get()  # B
                    sheet.cell(row=i, column=3).value = self.bc_var.get()  # C
                    sheet.cell(row=i, column=4).value = self.dia_var.get()  # D
                    sheet.cell(row=i, column=5).value = row_data[0]  # E
                    sheet.cell(row=i, column=6).value = row_data[1]  # F
                    sheet.cell(row=i, column=7).value = row_data[2]  # G
                    sheet.cell(row=i, column=8).value = "Low ADD"  # H
                # Set "High ADD" for original rows
                for i in range(2, num_rows + 2):
                    sheet.cell(row=i, column=8).value = "High ADD"  # H

            # Save file
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=self.product_name_var.get()
            )
            if save_path:
                wb.save(save_path)
                messagebox.showinfo("Success", f"File saved successfully at {save_path}")
            else:
                messagebox.showwarning("Warning", "Save operation cancelled")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
